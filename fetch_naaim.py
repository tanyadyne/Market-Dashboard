"""Fetch NAAIM Exposure Index full historical data and save as naaim.json.

Strategy:
  1. Fetch the NAAIM Exposure Index page HTML
  2. Extract the URL of the "Download EXCEL file with data since inception" link
  3. Download that Excel file (full history back to ~2006)
  4. Parse all rows, save to naaim.json sorted oldest -> newest

Output schema:
{
  "updated": "2026-05-05T07:00:00Z",
  "latest": 93.79,
  "q1_avg": 82.00,
  "rows": [
    {"date": "07/14/2006", "mean": 12.27, "bearish": -75, "q1": 0, "q2": 25, "q3": 75, "bullish": 200, "deviation": 65.21},
    ...
    {"date": "04/29/2026", "mean": 93.79, ...}
  ]
}
"""

import io
import json
import re
import sys
from datetime import datetime, timezone

try:
    import requests
except ImportError:
    print("ERROR: pip install requests", file=sys.stderr)
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("ERROR: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


PAGE_URL = "https://naaim.org/programs/naaim-exposure-index/"
OUTPUT = "naaim.json"

BROWSER_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}


def find_excel_url(html):
    """Find the most recent USE_Data-since-Inception_*.xlsx link in the page HTML."""
    matches = re.findall(
        r'https?://[^\s"\'<>]+USE_Data[-_]since[-_]Inception[^\s"\'<>]+\.xlsx',
        html, re.IGNORECASE)
    if not matches:
        return None
    # Prefer the most recently dated filename (lexicographic sort works for YYYY-MM-DD)
    matches.sort(reverse=True)
    return matches[0]


def parse_q1_avg(html):
    """Extract 'Last Quarter Average (Q1)' value from the page HTML."""
    m = re.search(
        r"Last\s+Quarter\s+Average\s*\([^)]*\)\s*(?:<[^>]*>\s*)*([\-0-9][0-9.]*)",
        html, re.IGNORECASE | re.DOTALL)
    return float(m.group(1)) if m else None


def parse_excel(content):
    """Parse the NAAIM Excel file, return list of row dicts sorted oldest -> newest."""
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = []
    header_seen = False
    for raw in ws.iter_rows(values_only=True):
        if raw is None:
            continue
        # Drop trailing None cells
        cells = list(raw)
        while cells and cells[-1] is None:
            cells.pop()
        if not cells:
            continue

        # Detect header row (contains "Date" or "NAAIM Number")
        first = str(cells[0]).strip() if cells[0] is not None else ""
        if not header_seen:
            joined = " ".join(str(c) for c in cells if c is not None).lower()
            if "date" in joined and ("naaim" in joined or "mean" in joined):
                header_seen = True
                continue
            # Skip pre-header junk rows
            continue

        # Need at least 8 cells: date, mean, bearish, q1, q2, q3, bullish, deviation
        if len(cells) < 8:
            continue

        # Normalize date: openpyxl may give a datetime, or a string like "07/14/2006"
        d = cells[0]
        if isinstance(d, datetime):
            date_str = d.strftime("%m/%d/%Y")
            sort_key = d
        else:
            ds = str(d).strip()
            if not ds or ds.lower() == "none":
                continue
            try:
                sort_key = datetime.strptime(ds, "%m/%d/%Y")
                date_str = sort_key.strftime("%m/%d/%Y")
            except ValueError:
                # Skip non-date rows (footers, totals, etc.)
                continue

        def num(x):
            if x is None or x == "":
                return None
            try:
                return float(x)
            except (TypeError, ValueError):
                return None

        rows.append({
            "_sort":     sort_key,
            "date":      date_str,
            "mean":      num(cells[1]),
            "bearish":   num(cells[2]),
            "q1":        num(cells[3]),
            "q2":        num(cells[4]),
            "q3":        num(cells[5]),
            "bullish":   num(cells[6]),
            "deviation": num(cells[7]),
        })

    # Dedupe by date (keep first occurrence), then sort oldest -> newest
    seen = set()
    uniq = []
    for r in rows:
        if r["date"] in seen:
            continue
        seen.add(r["date"])
        uniq.append(r)
    uniq.sort(key=lambda r: r["_sort"])
    for r in uniq:
        del r["_sort"]
    return uniq


def main():
    print(f"Fetching page: {PAGE_URL}")
    page = requests.get(PAGE_URL, headers=BROWSER_HEADERS, timeout=30)
    page.raise_for_status()
    html = page.text

    xlsx_url = find_excel_url(html)
    if not xlsx_url:
        print("ERROR: could not locate XLSX link in page HTML", file=sys.stderr)
        sys.exit(1)
    print(f"Found XLSX: {xlsx_url}")

    q1_avg = parse_q1_avg(html)
    print(f"Q1 average from page: {q1_avg}")

    xlsx_headers = dict(BROWSER_HEADERS)
    xlsx_headers["Referer"] = PAGE_URL
    xlsx_headers["Accept"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*"
    xlsx = requests.get(xlsx_url, headers=xlsx_headers, timeout=60)
    xlsx.raise_for_status()
    print(f"Downloaded {len(xlsx.content):,} bytes")

    rows = parse_excel(xlsx.content)
    if not rows:
        print("ERROR: no rows parsed from Excel", file=sys.stderr)
        sys.exit(1)
    print(f"Parsed {len(rows)} rows ({rows[0]['date']} -> {rows[-1]['date']})")

    latest = rows[-1]["mean"]

    out = {
        "updated": datetime.now(timezone.utc).isoformat(),
        "latest": latest,
        "q1_avg": q1_avg,
        "rows": rows,
    }

    with open(OUTPUT, "w") as f:
        json.dump(out, f, separators=(",", ":"))
    print(f"OK Saved {OUTPUT} (latest={latest}, q1_avg={q1_avg})")


if __name__ == "__main__":
    main()
